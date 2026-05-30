# -*- coding: utf-8 -*-
"""AUTO-EVO-AI V0.1 - Excel 引擎（A级）"""
# Grade: A
__module_meta__ = {"id":"excel-engine","name":"Excel Engine","version":"V0.1","group":"data","grade":"A",
    "tags":["data","excel","spreadsheet"],"description":"Excel引擎-真实Excel读写/CSV降级/格式检测"}

import csv, io, os, logging
from typing import Any, Dict, List, Optional, Union
from pathlib import Path
from modules._base.enterprise_module import (EnterpriseModule, ModuleStatus, HealthReport, CircuitBreakerMixin, RateLimiterMixin)

logger = logging.getLogger("evo.excel-engine")

# --- graceful degradation: 优先 openpyxl ---
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
    logger.info("excel_engine: openpyxl loaded, full Excel support enabled")
except ImportError:
    openpyxl = None
    HAS_OPENPYXL = False
    logger.warning("excel_engine: openpyxl not installed, falling back to CSV-only mode")


class ExcelEngine(CircuitBreakerMixin, RateLimiterMixin, EnterpriseModule):
    MODULE_ID = "excel-engine"
    MODULE_NAME = "Excel引擎"
    VERSION = "V0.1"
    MODULE_LEVEL = "A"

    def __init__(self, config=None):
        super().__init__(config)
        self._dataframes: Dict[str, Dict] = {}

    def initialize(self) -> None:
        self.status = ModuleStatus.RUNNING
        logger.info("ExcelEngine initialized (openpyxl=%s)", HAS_OPENPYXL)

    def health_check(self) -> HealthReport:
        return HealthReport(
            status=self.status.value, healthy=True, module_id=self.MODULE_ID,
            checks={"sheets": len(self._dataframes), "openpyxl": HAS_OPENPYXL}
        )

    async def execute(self, action, params=None):
        return await self._safe_execute(action, params, handler=self._dispatch)

    def _dispatch(self, p):
        a = p.get("action", "status")
        if a == "status":
            return {"success": True, "openpyxl": HAS_OPENPYXL, "sheets": len(self._dataframes)}
        if a == "read_excel":
            return self._action_read_excel(p)
        if a == "write_excel":
            return self._action_write_excel(p)
        if a == "detect_format":
            return self._action_detect_format(p)
        if a == "read_csv":
            return self._action_read_csv(p)
        if a == "write_csv":
            return self._action_write_csv(p)
        if a == "list":
            return {"success": True, "sheets": [
                {"id": k, "name": v.get("name", ""), "cols": len(v.get("headers", [])),
                 "rows": len(v.get("rows", []))}
                for k, v in self._dataframes.items()
            ], "count": len(self._dataframes)}
        if a == "delete":
            sid = p.get("sheet_id", "")
            self._dataframes.pop(sid, None)
            return {"success": True, "deleted": sid}
        if a == "stats":
            total_cells = sum(len(s.get("headers", [])) * len(s.get("rows", []))
                              for s in self._dataframes.values())
            return {"success": True, "total_sheets": len(self._dataframes),
                    "total_cells": total_cells,
                    "total_rows": sum(len(s.get("rows", [])) for s in self._dataframes.values())}
        return {"error": f"unknown:{a}"}

    # ── public API: read_excel ──────────────────────────────────────────
    def read_excel(self, path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """读取 Excel 或 CSV 文件，返回 list[dict]"""
        if not os.path.isfile(path):
            logger.error("read_excel: file not found: %s", path)
            return []
        ext = os.path.splitext(path)[1].lower()
        if ext in (".xlsx", ".xlsm", ".xls"):
            return self._read_xlsx(path, sheet_name)
        else:
            return self._read_csv_as_dicts(path)

    # ── public API: write_excel ─────────────────────────────────────────
    def write_excel(self, path: str, data: List[Dict[str, Any]],
                    sheet_name: str = "Sheet1") -> bool:
        """将 list[dict] 写入 Excel（优先 .xlsx，fallback .csv）"""
        if not data:
            logger.warning("write_excel: empty data, nothing written")
            return False
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx" and HAS_OPENPYXL:
            return self._write_xlsx(path, data, sheet_name)
        else:
            return self._write_csv(path, data)

    # ── public API: detect_format ───────────────────────────────────────
    def detect_format(self, path: str) -> Dict[str, Any]:
        """检测文件格式，返回格式信息"""
        if not os.path.isfile(path):
            return {"path": path, "exists": False}
        ext = os.path.splitext(path)[1].lower()
        size = os.path.getsize(path)
        result = {
            "path": path, "exists": True, "ext": ext,
            "size_bytes": size, "size_kb": round(size / 1024, 1)
        }
        if ext in (".xlsx", ".xlsm", ".xls"):
            result["type"] = "excel"
            if HAS_OPENPYXL:
                try:
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    result["sheet_count"] = len(wb.sheetnames)
                    result["sheet_names"] = wb.sheetnames
                    wb.close()
                except Exception as e:
                    logger.warning("detect_format: openpyxl read error: %s", e)
                    result["type"] = "excel (unreadable)"
                    result["error"] = str(e)
            else:
                result["type"] = "excel (openpyxl unavailable)"
        elif ext == ".csv":
            result["type"] = "csv"
            try:
                with open(path, "r", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    rows = sum(1 for _ in reader)
                result["row_count"] = rows
            except Exception as e:
                result["error"] = str(e)
        else:
            result["type"] = "unknown"
        return result

    # ── internal: action wrappers ───────────────────────────────────────
    def _action_read_excel(self, p):
        path = p.get("path", "")
        sheet = p.get("sheet_name")
        data = self.read_excel(path, sheet)
        return {"success": True, "path": path, "rows": len(data), "data": data[:100]}

    def _action_write_excel(self, p):
        path = p.get("path", "output.xlsx")
        data = p.get("data", [])
        sheet = p.get("sheet_name", "Sheet1")
        ok = self.write_excel(path, data, sheet)
        return {"success": ok, "path": path, "rows": len(data)}

    def _action_detect_format(self, p):
        path = p.get("path", "")
        return self.detect_format(path)

    def _action_read_csv(self, p):
        path = p.get("path", "")
        data = self._read_csv_as_dicts(path)
        return {"success": True, "path": path, "rows": len(data), "data": data[:100]}

    def _action_write_csv(self, p):
        path = p.get("path", "output.csv")
        data = p.get("data", [])
        ok = self._write_csv(path, data)
        return {"success": ok, "path": path, "rows": len(data)}

    # ── helpers: xlsx ───────────────────────────────────────────────────
    def _read_xlsx(self, path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        if not HAS_OPENPYXL:
            logger.warning("_read_xlsx: openpyxl unavailable, trying CSV fallback")
            return self._read_csv_as_dicts(path)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h) if h is not None else "" for h in next(rows_iter, [])]
            result = []
            for row in rows_iter:
                result.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
            wb.close()
            logger.info("_read_xlsx: loaded %d rows from %s", len(result), path)
            return result
        except Exception as e:
            logger.error("_read_xlsx: error reading %s: %s", path, e)
            return []

    def _write_xlsx(self, path: str, data: List[Dict[str, Any]],
                    sheet_name: str = "Sheet1") -> bool:
        if not HAS_OPENPYXL:
            logger.warning("_write_xlsx: openpyxl unavailable, falling back to CSV")
            return self._write_csv(path, data)
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([row.get(h, "") for h in headers])
            wb.save(path)
            wb.close()
            logger.info("_write_xlsx: wrote %d rows to %s", len(data), path)
            return True
        except Exception as e:
            logger.error("_write_xlsx: error writing %s: %s", path, e)
            return False

    # ── helpers: CSV ────────────────────────────────────────────────────
    def _read_csv_as_dicts(self, path: str) -> List[Dict[str, Any]]:
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                result = [dict(row) for row in reader]
            logger.info("_read_csv: loaded %d rows from %s", len(result), path)
            return result
        except Exception as e:
            logger.error("_read_csv: error reading %s: %s", path, e)
            return []

    def _write_csv(self, path: str, data: List[Dict[str, Any]]) -> bool:
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=list(data[0].keys()))
                writer.writeheader()
                writer.writerows(data)
            logger.info("_write_csv: wrote %d rows to %s", len(data), path)
            return True
        except Exception as e:
            logger.error("_write_csv: error writing %s: %s", path, e)
            return False

    async def shutdown(self) -> None:
        self._dataframes.clear()
        self.status = ModuleStatus.STOPPED
        logger.info("ExcelEngine shut down")


module_class = ExcelEngine
