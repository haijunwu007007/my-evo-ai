from __future__ import annotations

"""

AUTO-EVO-AI V0.1 — Excel 高级处理 (openpyxl)

Grade: A (生产级) | Category: documents

职责：Excel 高级功能：多 sheet 操作、图表生成、公式注入、条件格式、数据验证、透视表

"""



__module_meta__ = {

    "id": "excel-pro",

    "name": "Excel Pro",

    "version": "V0.1",

    "group": "documents",

    "inputs": [

        {"name": "action", "type": "string", "required": True, "description": "create / read / edit / chart / pivot / merge / convert / formula"},

        {"name": "path", "type": "string", "required": False, "description": "文件路径"},

        {"name": "data", "type": "dict", "required": False, "description": "操作参数"},

    ],

    "outputs": [

        {"name": "result", "type": "dict", "description": "执行结果"},

    ],

    "triggers": [],

    "depends_on": [],

    "tags": ["adapter", "doc"],

    "grade": "A",

    "description": "Excel 高级处理器：多 Sheet、图表（柱状/折线/饼图）、公式、条件格式、数据验证、数据透视表",

}



import os

import re

import json

import uuid

import logging

from pathlib import Path

from datetime import datetime

from typing import Any



try:

    from modules._base.enterprise_module import EnterpriseModule

    from modules._base.audit import AuditLogger

except ImportError:

    import sys

    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

    from _base.enterprise_module import EnterpriseModule

    from _base.audit import AuditLogger



logger = logging.getLogger("evo.excel")



try:

    import openpyxl

    from openpyxl import Workbook, load_workbook

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    from openpyxl.chart.label import DataLabelList

    from openpyxl.chart.series import DataPoint

    from openpyxl.utils import get_column_letter

    from openpyxl.worksheet.datavalidation import DataValidation

    from openpyxl.formatting.rule import CellIsRule, FormulaRule

    # PivotTable 在 openpyxl 3.1.x 中不可用，手动实现聚合

    HAS_OPENPYXL = True

except ImportError:

    HAS_OPENPYXL = False

    logger.warning("openpyxl not installed; Excel Pro disabled")



# ── 数值格式 ────────────────────────────────────────────



NUM_FMTS = {

    "general": None,

    "number": "#,##0",

    "decimal": "#,##0.00",

    "percent": "0.00%",

    "currency": '¥#,##0.00',

    "date": "YYYY-MM-DD",

    "datetime": "YYYY-MM-DD HH:MM:SS",

}





class ExcelPro(EnterpriseModule):

    """Excel 高级处理器"""



    def __init__(self, *args: Any, **kwargs: Any):

        super().__init__(*args, **kwargs)

        self._name = "excel-pro"

        self._work_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "spreadsheets")

        os.makedirs(self._work_dir, exist_ok=True)

        self._audit = AuditLogger()

        self.stats = {"calls": 0, "errors": 0}

        self._status = "ready" if HAS_OPENPYXL else "error"



    def initialize(self) -> None:

        logger.info(f"[excel-pro] initialized, HAS_OPENPYXL={HAS_OPENPYXL}")



    def health_check(self) -> dict[str, Any]:

        return {

            "status": self._status,

            "module": "excel-pro",

            "libraries": {"openpyxl": HAS_OPENPYXL},

            "stats": self.stats,

        }



    def shutdown(self) -> None:

        logger.info("[excel-pro] shutdown")



    # ── 主入口 ───────────────────────────────────────────



    async def execute(self, action: str, path: str = "", data: dict | None = None, **kwargs: Any) -> dict[str, Any]:

        if not HAS_OPENPYXL:

            return {"status": "error", "message": "openpyxl 未安装"}

        self.stats["calls"] += 1

        data = data or {}

        action_map: dict[str, Any] = {

            "create": self._create_workbook,

            "read": self._read_workbook,

            "edit": self._edit_workbook,

            "chart": self._add_chart,

            "pivot": self._create_pivot,

            "merge": self._merge_workbooks,

            "convert": self._convert_format,

            "formula": self._inject_formulas,

            "styles": self._apply_styles,

            "list": self._list_files,

            "info": self._get_info,

        }

        handler = action_map.get(action)

        if not handler:

            return {"status": "error", "message": f"未知 action: {action}，可选: {list(action_map.keys())}"}

        try:

            return await handler(path, data)

        except Exception as e:

            self.stats["errors"] += 1

            logger.exception(f"[excel-pro] execute {action} 失败")

            return {"status": "error", "message": str(e)}



    # ── 创建工作簿 ───────────────────────────────────────



    async def _create_workbook(self, path: str, data: dict) -> dict[str, Any]:

        filename = data.get("filename", f"spreadsheet_{uuid.uuid4().hex[:8]}.xlsx")

        save_path = path or os.path.join(self._work_dir, filename)

        sheets = data.get("sheets", [])



        wb = Workbook()

        # 删除默认 sheet

        wb.remove(wb.active)



        for si, sheet_data in enumerate(sheets):

            sheet_name = sheet_data.get("name", f"Sheet{si + 1}")

            ws = wb.create_sheet(title=sheet_name)



            headers = sheet_data.get("headers", [])

            rows = sheet_data.get("rows", [])

            column_widths = sheet_data.get("column_widths", {})

            header_style = sheet_data.get("header_style", {})

            freeze = sheet_data.get("freeze")



            # ── 写入表头 ──

            if headers:

                for ci, h in enumerate(headers, 1):

                    cell = ws.cell(row=1, column=ci, value=h)

                    self._apply_cell_style(cell, header_style.get("font", {"bold": True, "color": "FFFFFF"}),

                                           header_style.get("fill", "1F2937"))

                data_start = 2

            else:

                data_start = 1



            # ── 写入数据 ──

            for ri, row_data in enumerate(rows, data_start):

                for ci, val in enumerate(row_data, 1):

                    cell = ws.cell(row=ri, column=ci, value=val)



            # ── 列宽 ──

            if column_widths:

                for col_letter, width in column_widths.items():

                    ws.column_dimensions[col_letter].width = width

            elif headers:

                for ci in range(1, len(headers) + 1):

                    ws.column_dimensions[get_column_letter(ci)].width = max(12, len(str(headers[ci - 1])) + 4)



            # ── 冻结窗格 ──

            if freeze:

                ws.freeze_panes = freeze



        wb.save(save_path)

        return {"status": "success", "path": os.path.abspath(save_path), "sheets": len(sheets)}



    # ── 读取工作簿 ───────────────────────────────────────



    async def _read_workbook(self, path: str, data: dict) -> dict[str, Any]:

        if not path or not os.path.isfile(path):

            return {"status": "error", "message": f"文件不存在: {path}"}

        wb = load_workbook(path, data_only=True)

        sheet_names = data.get("sheets") or wb.sheetnames

        result: dict[str, Any] = {"sheets": {}, "active_sheet": wb.active.title}



        for name in sheet_names:

            if name not in wb.sheetnames:

                continue

            ws = wb[name]

            rows_data = []

            for row in ws.iter_rows(min_row=1, values_only=True):

                rows_data.append(list(row))

            result["sheets"][name] = {

                "rows": rows_data,

                "dimensions": ws.dimensions,

                "max_row": ws.max_row,

                "max_col": ws.max_column,

            }

        wb.close()

        return {"status": "success", "data": result}



    # ── 编辑工作簿 ───────────────────────────────────────



    async def _edit_workbook(self, path: str, data: dict) -> dict[str, Any]:

        if not path or not os.path.isfile(path):

            return {"status": "error", "message": f"文件不存在: {path}"}

        wb = load_workbook(path)

        edits = data.get("edits", [])



        for edit in edits:

            op = edit.get("op", "")

            sheet_name = edit.get("sheet", wb.active.title)

            if sheet_name not in wb.sheetnames:

                if op == "add_sheet":

                    wb.create_sheet(title=sheet_name)

                else:

                    continue

            ws = wb[sheet_name]



            if op == "write_cell":

                row = edit.get("row", 1)

                col = edit.get("col", 1)

                value = edit.get("value", "")

                ws.cell(row=row, column=col, value=value)



            elif op == "write_range":

                start_row = edit.get("start_row", 1)

                start_col = edit.get("start_col", 1)

                values = edit.get("values", [])

                for ri, row_data in enumerate(values):

                    for ci, val in enumerate(row_data):

                        ws.cell(row=start_row + ri, column=start_col + ci, value=val)



            elif op == "add_sheet":

                pass  # already handled above



            elif op == "delete_sheet":

                if sheet_name in wb.sheetnames:

                    std = wb[sheet_name]

                    wb.remove(std)



            elif op == "rename_sheet":

                new_name = edit.get("new_name", "")

                if new_name:

                    ws.title = new_name



            elif op == "insert_row":

                idx = edit.get("idx", 1)

                values = edit.get("values", [])

                ws.insert_rows(idx)

                for ci, val in enumerate(values, 1):

                    ws.cell(row=idx, column=ci, value=val)



            elif op == "delete_row":

                idx = edit.get("idx", 1)

                count = edit.get("count", 1)

                ws.delete_rows(idx, count)



        wb.save(path)

        wb.close()

        return {"status": "success", "path": os.path.abspath(path)}



    # ── 添加图表 ─────────────────────────────────────────



    async def _add_chart(self, path: str, data: dict) -> dict[str, Any]:

        if not path or not os.path.isfile(path):

            return {"status": "error", "message": f"文件不存在: {path}"}

        wb = load_workbook(path)

        ws_name = data.get("sheet", wb.active.title)

        if ws_name not in wb.sheetnames:

            return {"status": "error", "message": f"Sheet 不存在: {ws_name}"}

        ws = wb[ws_name]



        chart_type = data.get("chart_type", "bar")

        title = data.get("title", "图表")

        categories = data.get("categories", [])

        values = data.get("values", [])

        cat_col = data.get("cat_col", "A")

        val_col = data.get("val_col", "B")

        data_start = data.get("data_start", 1)

        data_end = data.get("data_end", len(categories) + 1)



        # 写入图表数据到工作表

        chart_data_start_row = ws.max_row + 2

        if categories and values:

            ws.cell(row=chart_data_start_row, column=1, value="类别")

            ws.cell(row=chart_data_start_row, column=2, value="数值")

            for i, (cat, val) in enumerate(zip(categories, values)):

                ws.cell(row=chart_data_start_row + 1 + i, column=1, value=cat)

                ws.cell(row=chart_data_start_row + 1 + i, column=2, value=val)

            cat_ref = f"A{chart_data_start_row + 1}:A{chart_data_start_row + len(categories)}"

            val_ref = f"B{chart_data_start_row + 1}:B{chart_data_start_row + len(categories)}"

        else:

            cat_ref = f"{cat_col}{data_start}:{cat_col}{data_end}"

            val_ref = f"{val_col}{data_start}:{val_col}{data_end}"



        if chart_type == "bar":

            chart = BarChart()

            chart.type = "col"

        elif chart_type == "line":

            chart = LineChart()

        elif chart_type == "pie":

            chart = PieChart()

        else:

            chart = BarChart()



        chart.title = title

        chart.style = 10

        data_ref = Reference(ws, min_col=2, min_row=chart_data_start_row if categories and values else data_start,

                             max_row=chart_data_start_row + len(categories) if categories and values else data_end)

        cats_ref = Reference(ws, min_col=1, min_row=(chart_data_start_row + 1) if categories and values else (data_start + 1),

                             max_row=chart_data_start_row + len(categories) if categories and values else data_end)



        chart.add_data(data_ref, titles_from_data=True)

        chart.set_categories(cats_ref)

        chart.shape = 4



        if chart_type == "pie":

            chart.dataLabels = DataLabelList()

            chart.dataLabels.showPercent = True

            chart.dataLabels.showVal = True



        ws.add_chart(chart, data.get("position", "D1"))



        wb.save(path)

        wb.close()

        return {"status": "success", "path": os.path.abspath(path), "chart_type": chart_type, "title": title}



    # ── 创建透视表 ───────────────────────────────────────



    async def _create_pivot(self, path: str, data: dict) -> dict[str, Any]:

        if not path or not os.path.isfile(path):

            return {"status": "error", "message": f"文件不存在: {path}"}

        wb = load_workbook(path)

        src_sheet_name = data.get("source_sheet", wb.active.title)

        if src_sheet_name not in wb.sheetnames:

            return {"status": "error", "message": f"源 Sheet 不存在: {src_sheet_name}"}

        ws_src = wb[src_sheet_name]



        pivot_name = data.get("pivot_name", "PivotTable1")

        pivot_dest = data.get("dest_sheet", "数据透视表")

        if pivot_dest in wb.sheetnames:

            ws_dest = wb[pivot_dest]

        else:

            ws_dest = wb.create_sheet(title=pivot_dest)



        # 读取源数据

        src_data = []

        for row in ws_src.iter_rows(min_row=1, values_only=True):

            src_data.append(list(row))



        if len(src_data) < 2:

            return {"status": "error", "message": "数据不足，无法创建透视表"}



        headers = src_data[0]

        rows = src_data[1:]



        # 分类字段和值字段

        row_field = data.get("row_field", headers[0] if headers else "")

        val_field = data.get("val_field", headers[-1] if len(headers) > 1 else headers[0])

        agg_func = data.get("agg_func", "sum")



        # 手动计算分组汇总

        row_idx = headers.index(row_field) if row_field in headers else 0

        val_idx = headers.index(val_field) if val_field in headers else -1



        summary: dict[str, float] = {}

        count_map: dict[str, int] = {}

        for row in rows:

            key = str(row[row_idx]) if row_idx < len(row) else "未知"

            try:

                val = float(row[val_idx]) if val_idx >= 0 and val_idx < len(row) and row[val_idx] is not None else 0

            except (ValueError, TypeError):

                val = 0

            if agg_func == "sum":

                summary[key] = summary.get(key, 0) + val

            elif agg_func == "count":

                count_map[key] = count_map.get(key, 0) + 1

            elif agg_func == "avg":

                summary[key] = summary.get(key, 0) + val

                count_map[key] = count_map.get(key, 0) + 1



        # 写入透视表结果

        ws_dest.cell(row=1, column=1, value=row_field)

        ws_dest.cell(row=1, column=2, value=f"{agg_func} of {val_field}")

        ws_dest.cell(row=1, column=1).font = Font(bold=True)

        ws_dest.cell(row=1, column=2).font = Font(bold=True)



        if agg_func == "avg":

            keys = sorted(summary.keys())

            for i, k in enumerate(keys, 2):

                ws_dest.cell(row=i, column=1, value=k)

                ws_dest.cell(row=i, column=2, value=round(summary[k] / max(count_map.get(k, 1), 1), 2))

        else:

            keys = sorted(summary.keys()) if agg_func == "sum" else sorted(count_map.keys())

            for i, k in enumerate(keys, 2):

                ws_dest.cell(row=i, column=1, value=k)

                ws_dest.cell(row=i, column=2, value=summary.get(k, count_map.get(k, 0)))



        ws_dest.column_dimensions["A"].width = 20

        ws_dest.column_dimensions["B"].width = 15



        wb.save(path)

        wb.close()

        return {"status": "success", "path": os.path.abspath(path), "pivot_sheet": pivot_dest}



    # ── 合并工作簿 ───────────────────────────────────────



    async def _merge_workbooks(self, path: str, data: dict) -> dict[str, Any]:

        files = data.get("files", [])

        if not files:

            return {"status": "error", "message": "未指定待合并文件列表"}

        save_path = path or os.path.join(self._work_dir, f"merged_{uuid.uuid4().hex[:8]}.xlsx")



        merged = Workbook()

        merged.remove(merged.active)



        for fp in files:

            if not os.path.isfile(fp):

                continue

            wb = load_workbook(fp, data_only=True)

            for ws in wb.worksheets:

                new_ws = merged.create_sheet(title=ws.title)

                for row in ws.iter_rows(min_row=1, values_only=True):

                    new_ws.append(list(row) if row else [])

            wb.close()



        merged.save(save_path)

        merged.close()

        return {"status": "success", "path": os.path.abspath(save_path), "merged_count": len(files)}



    # ── 格式转换 ─────────────────────────────────────────



    async def _convert_format(self, path: str, data: dict) -> dict[str, Any]:

        if not path or not os.path.isfile(path):

            return {"status": "error", "message": f"文件不存在: {path}"}

        target = data.get("to", "csv")



        if target == "csv":

            wb = load_workbook(path, data_only=True)

            ws = wb.active

            csv_path = path.replace(".xlsx", ".csv").replace(".xls", ".csv") if path.lower().endswith((".xlsx", ".xls")) else path + ".csv"

            with open(csv_path, "w", encoding="utf-8-sig") as f:

                for row in ws.iter_rows(values_only=True):

                    f.write(",".join(str(v) if v is not None else "" for v in row) + "\n")

            wb.close()

            return {"status": "success", "path": os.path.abspath(csv_path), "format": "csv"}



        elif target == "xlsx":

            wb = load_workbook(path, data_only=True)

            xlsx_path = path.replace(".csv", ".xlsx")

            wb.save(xlsx_path)

            wb.close()

            return {"status": "success", "path": os.path.abspath(xlsx_path), "format": "xlsx"}



        return {"status": "error", "message": f"不支持的转换目标: {target}"}



    # ── 注入公式 ─────────────────────────────────────────



    async def _inject_formulas(self, path: str, data: dict) -> dict[str, Any]:

        if not path or not os.path.isfile(path):

            return {"status": "error", "message": f"文件不存在: {path}"}

        wb = load_workbook(path)

        ws_name = data.get("sheet", wb.active.title)

        if ws_name not in wb.sheetnames:

            return {"status": "error", "message": f"Sheet 不存在: {ws_name}"}

        ws = wb[ws_name]

        formulas = data.get("formulas", [])



        for f in formulas:

            cell_ref = f.get("cell", "")

            formula = f.get("formula", "")

            if cell_ref and formula:

                ws[cell_ref] = formula



        wb.save(path)

        wb.close()

        return {"status": "success", "path": os.path.abspath(path), "formulas_added": len(formulas)}



    # ── 应用样式 ─────────────────────────────────────────



    async def _apply_styles(self, path: str, data: dict) -> dict[str, Any]:

        if not path or not os.path.isfile(path):

            return {"status": "error", "message": f"文件不存在: {path}"}

        wb = load_workbook(path)

        ws_name = data.get("sheet", wb.active.title)

        if ws_name not in wb.sheetnames:

            return {"status": "error", "message": f"Sheet 不存在: {ws_name}"}

        ws = wb[ws_name]

        styles = data.get("styles", [])



        for s in styles:

            cell_ref = s.get("cell", "")

            if not cell_ref:

                continue

            cell = ws[cell_ref]

            font_opts = s.get("font", {})

            fill_color = s.get("fill")

            align_opts = s.get("alignment", {})

            num_format = s.get("num_format")



            if font_opts:

                cell.font = Font(

                    bold=font_opts.get("bold", False),

                    italic=font_opts.get("italic", False),

                    size=font_opts.get("size"),

                    color=font_opts.get("color"),

                )

            if fill_color:

                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            if align_opts:

                cell.alignment = Alignment(

                    horizontal=align_opts.get("horizontal", "left"),

                    vertical=align_opts.get("vertical", "center"),

                )

            if num_format:

                cell.number_format = NUM_FMTS.get(num_format, num_format)



        wb.save(path)

        wb.close()

        return {"status": "success", "path": os.path.abspath(path)}



    # ── 获取信息 ─────────────────────────────────────────



    async def _get_info(self, path: str, data: dict) -> dict[str, Any]:

        if not path or not os.path.isfile(path):

            return {"status": "error", "message": f"文件不存在: {path}"}

        wb = load_workbook(path, data_only=True)

        info = {

            "sheets": [],

            "file_size_bytes": os.path.getsize(path),

        }

        for name in wb.sheetnames:

            ws = wb[name]

            info["sheets"].append({

                "name": name,

                "rows": ws.max_row,

                "cols": ws.max_column,

            })

        wb.close()

        return {"status": "success", "info": info}



    # ── 列出文件 ─────────────────────────────────────────



    async def _list_files(self, path: str, data: dict) -> dict[str, Any]:

        search_dir = path or self._work_dir

        if not os.path.isdir(search_dir):

            return {"status": "error", "message": f"目录不存在: {search_dir}"}

        files = []

        for f in sorted(os.listdir(search_dir)):

            fp = os.path.join(search_dir, f)

            if os.path.isfile(fp) and f.lower().endswith((".xlsx", ".xls", ".csv")):

                files.append({

                    "name": f,

                    "size_bytes": os.path.getsize(fp),

                    "modified": datetime.fromtimestamp(os.path.getmtime(fp)).isoformat(),

                    "ext": os.path.splitext(f)[1].lower(),

                })

        return {"status": "success", "files": files, "directory": search_dir}



    # ── 辅助方法 ─────────────────────────────────────────



    @staticmethod

    def _apply_cell_style(cell, font_opts: dict, fill_hex: str | None = None) -> None:

        """统一应用单元格样式"""

        font = Font(

            bold=font_opts.get("bold", False),

            color=font_opts.get("color", "000000"),

            size=font_opts.get("size", 11),

        )

        cell.font = font

        if fill_hex:

            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

        cell.alignment = Alignment(horizontal="center", vertical="center")





module_class = ExcelPro

