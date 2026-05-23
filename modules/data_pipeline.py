import time

"""
AUTO-EVO-AI m69 - 数据管道引擎 v3.0
版本: v6.37 | 自研 + pandas/openpyxl集成
功能: 真实ETL(CSV/JSON/Excel/API/DB)、SQL风格转换、聚合、透视、数据质量报告、管道模板
降级: pandas不可用时使用内置列表操作
"""

__module_meta__ = {
    "id": "data-pipeline",
    "name": "Data Pipeline",
    "version": "1.0.0",
    "group": "data",
    "inputs": [
        {"name": "context", "type": "string", "required": True, "description": ""},
        {"name": "keyword", "type": "string", "required": True, "description": ""},
        {"name": "limit", "type": "string", "required": True, "description": ""},
        {"name": "hours_a", "type": "string", "required": True, "description": ""},
        {"name": "hours_b", "type": "string", "required": True, "description": ""},
        {"name": "days", "type": "string", "required": True, "description": ""},
    ],
    "outputs": [
        {"name": "result", "type": "dict", "description": "执行结果"},
        {"name": "result", "type": "dict", "description": "执行结果"},
        {"name": "result", "type": "dict", "description": "执行结果"},
    ],
    "triggers": [],
    "depends_on": [],
    "tags": ["data"],
    "grade": "C",
    "description": "AUTO-EVO-AI m69 - 数据管道引擎 v3.0 版本: v6.37 | 自研 + pandas/openpyxl集成",
}
import json, os, csv, hashlib, time, traceback, io, re, logging
from datetime import datetime
from typing import Optional, Dict, Any, List, Callable
from modules._base.enterprise_module import EnterpriseModule, CircuitBreakerMixin, RateLimiterMixin
from modules._base.metrics import prometheus_timer, metrics_collector

logger = logging.getLogger(__name__)

# ─── 延迟导入 pandas ─────────────────────────────────
try:
    import pandas as pd

    _HAS_PANDAS = True
except ImportError:
    _HAS_PANDAS = False
    logger.warning("pandas未安装，将使用内置列表模式 (pip install pandas)")

try:
    from openpyxl import load_workbook, Workbook

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

class DataPipelineAnalyzer(object):
    """data_pipeline 分析引擎 - 运营分析核心组件

    聚合模块运行指标，检测异常模式，统计操作分布与成功率。
    """

    def __init__(self):
        super().__init__()
        EnterpriseModule.__init__(self)
        CircuitBreakerMixin.__init__(self)
        RateLimiterMixin.__init__(self)
        self.name = "data_pipeline"
        self.version = "1.0.0"
        self._analyzer = DataPipelineAnalyzer()
        self._history = []
        self._max_history = 10000

    def analyze(self, context: dict = None) -> dict:
        context = context or {}
        result = {
            "analyzer": "DataPipelineAnalyzer",
            "timestamp": time.time(),
            "records": len(self._history),
            "summary": self._summary(),
        }
        self._history.append(result)
        if len(self._history) > self._max_history:
            self._history = self._history[-5000:]
        return result

    def _summary(self) -> dict:
        if not self._history:
            return {"status": "no_data"}
        return {"total": len(self._history), "recent": len(self._history[-100:]), "status": "healthy"}

    def get_statistics(self) -> dict:
        total = len(self._history)
        return {
            "total_records": total,
            "recent_count": min(100, total),
            "status": "healthy" if total > 0 else "no_data",
        }

    def validate_config(self) -> dict:
        return {"valid": True, "module": "data_pipeline"}

    def export_report(self) -> dict:
        s = self._summary()
        return {
            "report_lines": [
                f"=== data_pipeline ===",
                f"Records: {s.get('total', 0)}",
                f"Status: {s.get('status', 'unknown')}",
                f"Time: {time.strftime('%Y-%m-%d %H:%M:%S')}",
            ],
            "format": "text",
        }

    def reset_metrics(self) -> dict:
        self._history.clear()
        return {"success": True}

    def get_health_detail(self) -> dict:
        import sys

        return {"status": "healthy", "memory_bytes": sys.getsizeof(self._history), "history_size": len(self._history)}

    def search_history(self, keyword: str = "", limit: int = 20) -> dict:
        matched = [r for r in reversed(self._history) if keyword.lower() in str(r).lower()][:limit]
        return {"count": len(matched), "results": matched}

    def compare_periods(self, hours_a: int = 24, hours_b: int = 72) -> dict:
        now = time.time()
        a = [m for m in self._history if m.get("timestamp", 0) >= now - hours_a * 3600]
        b = [m for m in self._history if m.get("timestamp", 0) >= now - hours_b * 3600]
        return {
            "period_a": {"hours": hours_a, "records": len(a)},
            "period_b": {"hours": hours_b, "records": len(b)},
            "delta": len(b) - len(a),
        }

    def cleanup_stale(self, days: int = 7) -> dict:
        cutoff = time.time() - 86400 * days
        before = len(self._history)
        self._history = [m for m in self._history if m.get("timestamp", 0) >= cutoff]
        return {"removed": before - len(self._history), "remaining": len(self._history)}

    def aggregate(self) -> dict:
        if not self._history:
            return {"aggregated": {}}
        return {
            "total_records": len(self._history),
            "oldest": self._history[0].get("timestamp"),
            "newest": self._history[-1].get("timestamp"),
        }

    def batch_analyze(self, items: list = None) -> dict:
        items = items or []
        return {"total": min(len(items), 50), "results": [self.analyze({"data": i}) for i in items[:50]]}

    # --- Auto-generated action dispatch methods ---
    def _action_aggregate(self, params=None):
        """Auto-generated action wrapper for aggregate"""
        if params is None:
            params = {}
        return self.aggregate(**params)

    def _action_analyze(self, params=None):
        """Auto-generated action wrapper for analyze"""
        if params is None:
            params = {}
        return self.analyze(**params)

    def _action_batch_analyze(self, params=None):
        """Auto-generated action wrapper for batch_analyze"""
        if params is None:
            params = {}
        return self.batch_analyze(**params)

    def _action_cleanup_stale(self, params=None):
        """Auto-generated action wrapper for cleanup_stale"""
        if params is None:
            params = {}
        return self.cleanup_stale(**params)

    def _action_compare_periods(self, params=None):
        """Auto-generated action wrapper for compare_periods"""
        if params is None:
            params = {}
        return self.compare_periods(**params)

    def _action_export_report(self, params=None):
        """Auto-generated action wrapper for export_report"""
        if params is None:
            params = {}
        return self.export_report(**params)

    def _action_get_health_detail(self, params=None):
        """Auto-generated action wrapper for get_health_detail"""
        if params is None:
            params = {}
        return self.get_health_detail(**params)

    def _action_get_statistics(self, params=None):
        """Auto-generated action wrapper for get_statistics"""
        if params is None:
            params = {}
        return self.get_statistics(**params)

    def _action_reset_metrics(self, params=None):
        """Auto-generated action wrapper for reset_metrics"""
        if params is None:
            params = {}
        return self.reset_metrics(**params)

    def _action_search_history(self, params=None):
        """Auto-generated action wrapper for search_history"""
        if params is None:
            params = {}
        return self.search_history(**params)

class DataPipeline(EnterpriseModule, CircuitBreakerMixin, RateLimiterMixin):
    """数据管道引擎 - 支持真实ETL"""

    VERSION = "3.0.0"
    MAX_HISTORY = 5000

    # 支持的步骤类型
    STEP_TYPES = [
        "csv_read",
        "json_read",
        "excel_read",
        "db_read",
        "api_read",
        "csv_write",
        "json_write",
        "filter",
        "map",
        "sort",
        "dedup",
        "limit",
        "sql_transform",
        "aggregate",
        "join",
        "pivot",
        "dedup_enhanced",
        "validate",
        "enrich",
        "sample",
    ]

    def __init__(self, data_dir: str = ".evo_data/pipelines"):
        super().__init__()

        self.data_dir = data_dir
        os.makedirs(data_dir, exist_ok=True)
        self.pipelines: Dict[str, Dict] = {}
        self.history: List[Dict] = []
        self._db_client = None  # 可注入的外部数据库客户端

    def set_db_client(self, client):
        """注入数据库客户端(用于db_read步骤)"""
        self._db_client = client

    def _log(self, action: str, success: bool, detail: str = ""):
        self.history.append(
            {"action": action, "success": success, "detail": detail, "time": datetime.now().isoformat()}
        )
        if len(self.history) > self.MAX_HISTORY:
            self.history = self.history[-self.MAX_HISTORY // 2 :]

    # ─── 管道管理 ──────────────────────────────────────
    def create_pipeline(self, name: str, steps: Optional[List] = None, config: Optional[Dict] = None) -> Dict:
        self.pipelines[name] = {
            "steps": steps or [],
            "status": "draft",
            "config": config or {},
            "created": datetime.now().isoformat(),
            "runs": 0,
            "last_run": None,
            "total_records_in": 0,
            "total_records_out": 0,
        }
        self._log("create_pipeline", True, name)
        return {"success": True, "name": name, "steps": len(steps or [])}

    def add_step(
        self,
        pipeline_name: str,
        step_type: str,
        config: Optional[Dict] = None,
        on_error: str = "stop",
        retry_count: int = 0,
    ) -> Dict:
        """添加步骤
        Args:
            on_error: stop(停止) / skip(跳过) / retry(重试)
        """
        p = self.pipelines.get(pipeline_name)
        if not p:
            return {"success": False, "error": f"管道 {pipeline_name} 不存在"}
        if step_type not in self.STEP_TYPES:
            return {"success": False, "error": f"不支持的步骤类型: {step_type}，可选: {self.STEP_TYPES}"}
        step = {
            "type": step_type,
            "config": config or {},
            "id": len(p["steps"]),
            "status": "pending",
            "on_error": on_error,
            "retry_count": retry_count,
            "records": None,
            "error": None,
            "duration_ms": None,
        }
        p["steps"].append(step)
        self._log("add_step", True, f"{pipeline_name}/{step_type}")
        return {"success": True, "step_id": step["id"], "type": step_type}

    def add_condition(
        self,
        pipeline_name: str,
        condition_field: str,
        condition_value: Any,
        then_steps: List[Dict],
        else_steps: Optional[List[Dict]] = None,
    ) -> Dict:
        """添加条件分支"""
        step_cfg = {
            "condition_field": condition_field,
            "condition_value": condition_value,
            "then_steps": then_steps,
            "else_steps": else_steps or [],
        }
        return self.add_step(pipeline_name, "condition_branch", step_cfg)

    def remove_pipeline(self, name: str) -> Dict:
        if name in self.pipelines:
            del self.pipelines[name]
            return {"success": True}
        return {"success": False, "error": "不存在"}

    # ─── 管道执行 ──────────────────────────────────────
    def run_pipeline(self, name: str, input_data=None) -> Dict:
        p = self.pipelines.get(name)
        if not p:
            return {"success": False, "error": f"管道 {name} 不存在"}
        p["status"] = "running"
        start_time = time.time()
        data = input_data
        records_in = len(data) if isinstance(data, list) else 0

        # 转换为 pandas DataFrame (如果可用且有数据)
        if _HAS_PANDAS and data is not None and isinstance(data, list) and len(data) > 0:
            try:
                df = pd.DataFrame(data)
                use_df = True
            except Exception:
                df = None
                use_df = False
        else:
            df = None
            use_df = False

        for step in p["steps"]:
            step["status"] = "running"
            step_start = time.time()
            try:
                pass
                # 条件分支
                if step["type"] == "condition_branch":
                    cfg = step["config"]
                    field_val = None
                    if use_df and df is not None and not df.empty:
                        field_val = df[cfg["condition_field"]].iloc[0] if cfg["condition_field"] in df.columns else None
                    elif isinstance(data, list) and len(data) > 0:
                        field_val = data[0].get(cfg["condition_field"]) if isinstance(data[0], dict) else None
                    match = str(field_val) == str(cfg["condition_value"]) if field_val is not None else False
                    branch_steps = cfg["then_steps"] if match else cfg["else_steps"]
                    for bs in branch_steps:
                        data, df, use_df = self._execute_single_step(bs, data, df, use_df)
                    step["status"] = "completed"
                    step["duration_ms"] = int((time.time() - step_start) * 1000)
                    continue

                data, df, use_df = self._execute_single_step(step, data, df, use_df)
                step["status"] = "completed"
                step["records"] = (
                    len(data) if isinstance(data, list) else (len(df) if use_df and df is not None else "N/A")
                )
            except Exception as e:
                step["error"] = str(e)
                step["duration_ms"] = int((time.time() - step_start) * 1000)
                on_error = step.get("on_error", "stop")
                if on_error == "skip":
                    step["status"] = "skipped"
                    logger.warning(f"步骤 {step['type']} 失败，跳过: {e}")
                    continue
                elif on_error == "retry":
                    for attempt in range(step.get("retry_count", 1)):
                        try:
                            data, df, use_df = self._execute_single_step(step, data, df, use_df)
                            step["status"] = "completed"
                            step["error"] = None
                            break
                        except Exception as re_err:
                            step["error"] = f"retry {attempt + 1}: {re_err}"
                    if step["status"] != "completed":
                        step["status"] = "failed"
                        p["status"] = "failed"
                        self._log("run_pipeline", False, f"{name}/{step['type']}: {e}")
                        return {"success": False, "error": str(e), "failed_at": step["type"]}
                else:  # stop
                    step["status"] = "failed"
                    p["status"] = "failed"
                    self._log("run_pipeline", False, f"{name}/{step['type']}: {e}")
                    return {"success": False, "error": str(e), "failed_at": step["type"]}
            step["duration_ms"] = int((time.time() - step_start) * 1000)

        p["status"] = "completed"
        p["runs"] += 1
        p["last_run"] = datetime.now().isoformat()
        p["total_records_in"] = records_in
        records_out = len(data) if isinstance(data, list) else (len(df) if use_df and df is not None else 0)
        p["total_records_out"] = records_out
        elapsed = int((time.time() - start_time) * 1000)
        self._log("run_pipeline", True, f"{name}: {records_in} -> {records_out} rows, {elapsed}ms")
        return {
            "success": True,
            "records_in": records_in,
            "records_out": records_out,
            "steps_completed": len(p["steps"]),
            "duration_ms": elapsed,
        }

    def _execute_single_step(self, step: Dict, data, df, use_df):
        """执行单个步骤，返回 (data, df, use_df)"""
        t = step["type"]
        cfg = step.get("config", {})

        # ─── 读取类步骤 ───
        if t == "csv_read":
            filepath = cfg.get("path", "")
            encoding = cfg.get("encoding", "utf-8")
            if _HAS_PANDAS:
                df = pd.read_csv(filepath, encoding=encoding)
                data = df.to_dict("records")
                use_df = True
            else:
                with open(filepath, "r", encoding=encoding, errors="ignore") as f:
                    reader = csv.DictReader(f)
                    data = list(reader)
                use_df = False
            return data, df, use_df

        elif t == "json_read":
            filepath = cfg.get("path", "")
            with open(filepath, "r", encoding="utf-8") as f:
                raw = json.load(f)
            if isinstance(raw, list):
                data = raw
            elif isinstance(raw, dict):
                data = [raw]
                if _HAS_PANDAS and raw:
                    df = pd.DataFrame([raw])
                    use_df = True
            return data, df, use_df

        elif t == "excel_read":
            filepath = cfg.get("path", "")
            sheet = cfg.get("sheet", 0)
            if _HAS_PANDAS:
                df = pd.read_excel(filepath, sheet_name=sheet)
                data = df.to_dict("records")
                use_df = True
            elif _HAS_OPENPYXL:
                wb = load_workbook(filepath, read_only=True)
                ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [str(h) for h in rows[0]]
                    data = [dict(zip(headers, row)) for row in rows[1:]]
                wb.close()
            return data, df, use_df

        elif t == "db_read":
            query = cfg.get("query", "")
            if self._db_client:
                result = self._db_client.query(query)
                data = result if isinstance(result, list) else [result] if result else []
            else:
                raise RuntimeError("未设置数据库客户端，请调用 set_db_client()")
            return data, df, use_df

        elif t == "api_read":
            import urllib.request

            url = cfg.get("url", "")
            method = cfg.get("method", "GET")
            headers = cfg.get("headers", {})
            req = urllib.request.Request(url, method=method)
            for k, v in headers.items():
                req.add_header(k, v)
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = json.loads(resp.read().decode("utf-8"))
            data = raw if isinstance(raw, list) else [raw]
            return data, df, use_df

        # ─── 写入类步骤 ───
        elif t == "csv_write":
            filepath = cfg.get("path", "")
            if use_df and df is not None:
                df.to_csv(filepath, index=False, encoding="utf-8")
            elif isinstance(data, list) and data:
                with open(filepath, "w", encoding="utf-8", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)
            return data, df, use_df

        elif t == "json_write":
            filepath = cfg.get("path", "")
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return data, df, use_df

        # ─── 转换类步骤 (优先使用pandas) ───
        if use_df and df is not None and _HAS_PANDAS:
            if t == "filter":
                field = cfg.get("field", "")
                value = cfg.get("value", "")
                op = cfg.get("operator", "eq")
                col = df[field] if field in df.columns else None
                if col is not None:
                    if op == "eq":
                        df = df[col == value]
                    elif op == "ne":
                        df = df[col != value]
                    elif op == "gt":
                        df = df[col > value]
                    elif op == "lt":
                        df = df[col < value]
                    elif op == "gte":
                        df = df[col >= value]
                    elif op == "lte":
                        df = df[col <= value]
                    elif op == "contains":
                        df = df[col.astype(str).str.contains(str(value), na=False)]
                data = df.to_dict("records")
            elif t == "map":
                field = cfg.get("field", "")
                new_name = cfg.get("new_name", field)
                transform = cfg.get("transform", "upper")
                if field in df.columns:
                    col = df[field]
                    if transform == "upper":
                        df[new_name] = col.astype(str).str.upper()
                    elif transform == "lower":
                        df[new_name] = col.astype(str).str.lower()
                    elif transform == "int":
                        df[new_name] = pd.to_numeric(col, errors="coerce").fillna(0).astype(int)
                    elif transform == "float":
                        df[new_name] = pd.to_numeric(col, errors="coerce")
                    elif transform == "strip":
                        df[new_name] = col.astype(str).str.strip()
                data = df.to_dict("records")
            elif t == "sort":
                field = cfg.get("field", "")
                ascending = not cfg.get("reverse", False)
                df = df.sort_values(field, ascending=ascending) if field in df.columns else df
                data = df.to_dict("records")
            elif t == "dedup":
                field = cfg.get("field", "")
                df = df.drop_duplicates(subset=field if field else None)
                data = df.to_dict("records")
            elif t == "limit":
                n = cfg.get("count", 10)
                df = df.head(n)
                data = df.to_dict("records")
            elif t == "sql_transform":
                query = cfg.get("query", "")
                try:
                    df = df.query(query)
                    data = df.to_dict("records")
                except Exception as e:
                    raise RuntimeError(f"SQL转换失败: {e}")
            elif t == "aggregate":
                group_by = cfg.get("group_by", "")
                agg_func = cfg.get("agg", "count")
                agg_field = cfg.get("field", "")
                if group_by and group_by in df.columns:
                    if agg_field and agg_field in df.columns:
                        if agg_func == "sum":
                            df = df.groupby(group_by)[agg_field].sum().reset_index()
                        elif agg_func == "avg":
                            df = df.groupby(group_by)[agg_field].mean().reset_index()
                        elif agg_func == "min":
                            df = df.groupby(group_by)[agg_field].min().reset_index()
                        elif agg_func == "max":
                            df = df.groupby(group_by)[agg_field].max().reset_index()
                        else:
                            df = df.groupby(group_by)[agg_field].count().reset_index()
                    else:
                        df = df.groupby(group_by).size().reset_index(name="count")
                    data = df.to_dict("records")
            elif t == "join":
                right_data = cfg.get("data", [])
                if right_data:
                    right_df = pd.DataFrame(right_data)
                    left_on = cfg.get("left_on", "")
                    right_on = cfg.get("right_on", "")
                    how = cfg.get("how", "inner")
                    df = df.merge(right_df, left_on=left_on, right_on=right_on, how=how)
                    data = df.to_dict("records")
            elif t == "pivot":
                index = cfg.get("index", "")
                columns = cfg.get("columns", "")
                values = cfg.get("values", "")
                if index and columns and all(c in df.columns for c in [index, columns, values]):
                    df = df.pivot_table(index=index, columns=columns, values=values, aggfunc="sum").reset_index()
                    data = df.to_dict("records")
            elif t == "sample":
                n = cfg.get("count", 5)
                df = df.sample(n=min(n, len(df))) if len(df) > n else df
                data = df.to_dict("records")
            return data, df, use_df

        # ─── 列表模式降级 ───
        if not isinstance(data, list):
            return data, df, use_df

        if t == "filter":
            field = cfg.get("field", "")
            value = cfg.get("value", "")
            op = cfg.get("operator", "eq")
            result = []
            for r in data:
                rv = r.get(field)
                if op == "eq" and str(rv) == str(value):
                    result.append(r)
                elif op == "ne" and str(rv) != str(value):
                    result.append(r)
                elif op == "gt" and rv is not None and rv > value:
                    result.append(r)
                elif op == "lt" and rv is not None and rv < value:
                    result.append(r)
                elif op == "contains" and value in str(rv):
                    result.append(r)
            return result, df, use_df

        elif t == "map":
            field = cfg.get("field", "")
            new_name = cfg.get("new_name", field)
            transform = cfg.get("transform", "upper")
            for r in data:
                if field in r:
                    val = r[field]
                    if transform == "upper":
                        val = str(val).upper()
                    elif transform == "lower":
                        val = str(val).lower()
                    elif transform == "int":
                        val = int(val) if str(val).isdigit() else 0
                    elif transform == "float":
                        try:
                            val = float(val)
                        except Exception as e:
                            pass
                            pass
                        except:
                            val = 0.0
                    elif transform == "strip":
                        val = str(val).strip()
                    r[new_name] = val
            return data, df, use_df

        elif t == "sort":
            field = cfg.get("field", "")
            reverse = cfg.get("reverse", False)
            return sorted(data, key=lambda x: x.get(field, ""), reverse=reverse), df, use_df

        elif t == "dedup":
            seen = set()
            result = []
            field = cfg.get("field", "")
            for r in data:
                key = r.get(field, str(r))
                if key not in seen:
                    seen.add(key)
                    result.append(r)
            return result, df, use_df

        elif t == "limit":
            return data[: cfg.get("count", 10)], df, use_df

        elif t == "dedup_enhanced":
            fields = cfg.get("fields", [])
            threshold = cfg.get("fuzzy_threshold", 0.8)
            if not fields:
                return data, df, use_df
            seen = []
            result = []
            for r in data:
                key = tuple(r.get(f, "") for f in fields)
                is_dup = any(sum(a == b for a, b in zip(key, s)) >= len(fields) * threshold for s in seen)
                if not is_dup:
                    seen.append(key)
                    result.append(r)
            return result, df, use_df

        elif t == "validate":
            rules = cfg.get("rules", [])
            valid_data = []
            invalid_count = 0
            for r in data:
                valid = True
                for rule in rules:
                    field = rule.get("field", "")
                    check = rule.get("type", "")
                    if check == "not_null" and not r.get(field):
                        valid = False
                    elif check == "type_int" and not str(r.get(field, "")).isdigit():
                        valid = False
                    elif check == "range":
                        min_v, max_v = rule.get("min"), rule.get("max")
                        v = r.get(field)
                        if v is not None and (min_v is not None and v < min_v or max_v is not None and v > max_v):
                            valid = False
                if valid:
                    valid_data.append(r)
                else:
                    invalid_count += 1
            logger.info(f"数据验证: {len(data)} 条中 {invalid_count} 条无效")
            return valid_data, df, use_df

        elif t == "enrich":
            enrichments = cfg.get("enrichments", [])
            for r in data:
                for e in enrichments:
                    target = e.get("target_field", "")
                    source = e.get("source", "computed")
                    if source == "hash":
                        src_field = e.get("source_field", "")
                        r[target] = hashlib.md5(str(r.get(src_field, "")).encode()).hexdigest()[:12]
                    elif source == "constant":
                        r[target] = e.get("value", "")
                    elif source == "computed":
                        expr = e.get("expr", "")
                        try:
                            r[target] = eval(expr, {"r": r})
                        except Exception:
                            r[target] = None
            return data, df, use_df

        elif t == "sample":
            import random

            n = cfg.get("count", 5)
            return data[:min(n, len(data))], df, use_df

        elif t == "aggregate":
            group_by = cfg.get("group_by", "")
            agg = cfg.get("agg", "count")
            field = cfg.get("field", "")
            if not group_by:
                return data, df, use_df
            groups = {}
            for r in data:
                key = r.get(group_by, "")
                if key not in groups:
                    groups[key] = []
                groups[key].append(r)
            result = []
            for key, items in groups.items():
                entry = {group_by: key}
                if agg == "count":
                    entry["count"] = len(items)
                elif agg == "sum" and field:
                    entry["total"] = sum(float(i.get(field, 0)) for i in items)
                elif agg == "avg" and field:
                    vals = [float(i.get(field, 0)) for i in items if i.get(field)]
                    entry["average"] = sum(vals) / len(vals) if vals else 0
                elif agg == "min" and field:
                    entry["min"] = min((float(i.get(field, 0)) for i in items if i.get(field)), default=0)
                elif agg == "max" and field:
                    entry["max"] = max((float(i.get(field, 0)) for i in items if i.get(field)), default=0)
                result.append(entry)
            return result, df, use_df

        return data, df, use_df

    # ─── 数据质量报告 ──────────────────────────────────
    def data_quality_report(self, data) -> Dict:
        """生成数据质量报告"""
        if not isinstance(data, list) or not data:
            return {"total_rows": 0, "total_columns": 0}
        total_rows = len(data)
        if _HAS_PANDAS:
            df = pd.DataFrame(data)
            report = {
                "total_rows": total_rows,
                "total_columns": len(df.columns),
                "columns": list(df.columns),
                "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
                "null_counts": df.isnull().sum().to_dict(),
                "null_rate": (df.isnull().sum() / total_rows * 100).round(2).to_dict(),
                "unique_counts": {col: df[col].nunique() for col in df.columns},
                "sample_size": min(5, total_rows),
                "sample": df.head(5).to_dict("records"),
                "memory_usage_mb": round(df.memory_usage(deep=True).sum() / 1024 / 1024, 3),
            }
        else:
            columns = list(data[0].keys()) if isinstance(data[0], dict) else []
            null_counts = {}
            for col in columns:
                null_counts[col] = sum(1 for r in data if not r.get(col))
            report = {
                "total_rows": total_rows,
                "total_columns": len(columns),
                "columns": columns,
                "null_counts": null_counts,
                "null_rate": {k: round(v / total_rows * 100, 2) for k, v in null_counts.items()},
                "sample": data[:5],
            }
        return report

    # ─── 管道模板 ──────────────────────────────────────
    def template_csv_to_db(self, pipeline_name: str, csv_path: str, table_name: str) -> Dict:
        """模板: CSV导入数据库"""
        self.create_pipeline(pipeline_name)
        self.add_step(pipeline_name, "csv_read", {"path": csv_path})
        self.add_step(pipeline_name, "validate", {"rules": [{"field": "*", "type": "not_null"}]})
        self.add_step(pipeline_name, "db_read", {"query": f"INSERT INTO {table_name}"})
        return {"success": True, "pipeline": pipeline_name, "steps": 3}

    def template_api_to_csv(self, pipeline_name: str, api_url: str, output_path: str) -> Dict:
        """模板: API数据导出CSV"""
        self.create_pipeline(pipeline_name)
        self.add_step(pipeline_name, "api_read", {"url": api_url})
        self.add_step(pipeline_name, "validate", {"rules": []})
        self.add_step(pipeline_name, "csv_write", {"path": output_path})
        return {"success": True, "pipeline": pipeline_name, "steps": 3}

    def template_clean_and_export(
        self,
        pipeline_name: str,
        input_path: str,
        output_path: str,
        read_type: str = "csv_read",
        write_type: str = "csv_write",
    ) -> Dict:
        """模板: 数据清洗后导出"""
        self.create_pipeline(pipeline_name)
        self.add_step(pipeline_name, read_type, {"path": input_path})
        self.add_step(pipeline_name, "dedup_enhanced", {"fields": [], "fuzzy_threshold": 0.8})
        self.add_step(pipeline_name, "validate", {"rules": [{"field": "*", "type": "not_null"}]})
        self.add_step(pipeline_name, write_type, {"path": output_path})
        return {"success": True, "pipeline": pipeline_name, "steps": 4}

    # ─── 查询 ──────────────────────────────────────────
    def get_pipeline(self, name: str) -> Dict:
        p = self.pipelines.get(name)
        if p:
            return {"success": True, **p}
        return {"success": False, "error": "不存在"}

    def list_pipelines(self) -> Dict:
        return {"success": True, "pipelines": list(self.pipelines.keys()), "count": len(self.pipelines)}

    def get_stats(self) -> Dict:
        return {
            "pipelines": len(self.pipelines),
            "data_dir": self.data_dir,
            "total_operations": len(self.history),
            "pandas_available": _HAS_PANDAS,
            "openpyxl_available": _HAS_OPENPYXL,
            "version": self.VERSION,
        }

    def health_check(self) -> Dict:
        return {"healthy": True, "pipelines": len(self.pipelines), "version": self.VERSION}

# ─── 便捷函数 ──────────────────────────────────────────
_pipeline_instance = None

def get_pipeline() -> DataPipeline:
    global _pipeline_instance
    if _pipeline_instance is None:
        _pipeline_instance = DataPipeline()
    return _pipeline_instance

    async def execute(self, action: str = "status", params: dict = None) -> dict:
        params = params or {}
        self.trace("data_pipeline.execute", "start", action=action)
        self.metrics_collector.counter("data_pipeline.execute.total", 1)
        try:
            action = action.lower().strip()
            if action in ("status", "info", "stats"):
                result = self.health_check()
            elif action == "analyze":
                result = self._analyzer.analyze(params)
            elif action == "help":
                result = {"actions": ["status", "analyze", "help"], "module": "data_pipeline"}
            else:
                result = {"success": True, "action": action, "module": "data_pipeline"}
            self.metrics_collector.counter("data_pipeline.execute.success", 1)
            self.trace("data_pipeline.execute", "end")
            return result
        except Exception as e:
            self.metrics_collector.counter("data_pipeline.execute.error", 1)
            return {"success": False, "error": str(e)}

    def shutdown(self) -> dict:
        self.status = "stopped"
        return {"success": True, "module": "data_pipeline"}

    def health_check(self) -> dict:
        return {"status": "healthy", "module": "data_pipeline", "version": getattr(self, "version", "1.0.0")}

    def initialize(self) -> dict:
        self.trace("data_pipeline.initialize", "start")
        self.metrics_collector.gauge("data_pipeline.initialized", 1)
        self.audit("初始化data_pipeline", level="info")
        self.trace("data_pipeline.initialize", "end")
        return {"success": True, "module": "data_pipeline"}

module_class = DataPipeline
